from flask import Flask, render_template, request, jsonify
import os
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import hashlib

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}
app.config['EXCEL_FILE'] = 'crime_data.xlsx'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def save_to_excel(name, crime, image_path):
    excel_file = app.config['EXCEL_FILE']

    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Crime'
        sheet['C1'] = 'Image Path'
        workbook.save(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=crime)
    sheet.cell(row=row, column=3, value=image_path)
    
    img = ExcelImage(image_path)
    img.width = 60
    img.height = 80
    sheet.add_image(img, 'D{}'.format(row))

    workbook.save(excel_file)

def hash_file(file_path):
    hasher = hashlib.sha256()
    with open(file_path, 'rb') as file:
        while chunk := file.read(8192):
            hasher.update(chunk)
    return hasher.hexdigest()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'})

    file = request.files['file']
    name = request.form['name']
    crime = request.form['crime']

    if file.filename == '':
        return jsonify({'message': 'No selected file'})

    if file and allowed_file(file.filename):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        save_to_excel(name, crime, file_path)

        return jsonify({'message': 'Data uploaded successfully'})
    else:
        return jsonify({'message': 'Invalid file format'})

@app.route('/retrieve', methods=['POST'])
def retrieve():
    data = request.get_json()
    image_url = data.get('imageUrl', '')

    image_path = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(image_url))

    excel_file = app.config['EXCEL_FILE']
    
    if not os.path.exists(excel_file):
        return jsonify({'message': 'Data not found'})

    image_hash = hash_file(image_path)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        cell_hash = hash_file(cell_value)
        print(f"Checking: {image_hash} == {cell_hash}")
        if cell_hash and cell_hash == image_hash:
            name = sheet.cell(row=row, column=1).value
            crime = sheet.cell(row=row, column=2).value
            return jsonify({'message': f'Name: {name}, Crime: {crime}'})
    
    return jsonify({'message': 'Data not found'})

if __name__ == '__main__':
    app.run(debug=True)
